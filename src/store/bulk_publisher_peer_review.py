import json
import os
import re
from store.bulk import dump_file
from configparser import ConfigParser
from datetime import datetime
from glob import glob
from openpyxl import load_workbook
from utils.utils import ROOT_DIR


def clean_val(val, mapping={}):
    if type(val) == datetime:
        return str(val.date())
    if not type(val) == str:
        if val == 1:
            return True
        elif val == 0:
            return False
        return val
    if mapping.get(val.lower()):
        return { "@id": mapping.get(val.lower()) }
    return val


def clean_key(key, mapping={}):
    if not key:
        key = "None"
    key = str(key).lower()
    if mapping.get(key): return mapping.get(key)
    key = re.sub("[-/ ]", "_", key)
    key = re.sub("[()]", "", key)
    key = re.sub("_([a-z])", lambda c: c.group(1).upper(), key)
    return key


def load_publishers_excel_file(file, mapping={}):
    provenance_rows = [1, 2, 3]
    category_row = 4
    property_row = 5
    first_data_row = 6
    
    workbook = load_workbook(file)
    worksheet = workbook.active
    result = []
    for row in range(first_data_row, worksheet.max_row):
        record = {}
        # provenance
        for r in provenance_rows:
            key = clean_key(worksheet[r][0].value, mapping)
            val = clean_val(worksheet[r][1].value, mapping)
            if not record.get("provenance"):
                record["provenance"] = {}
            record["provenance"][key] = val
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if col[row].value == None:
                continue
            cat = clean_key(col[category_row].value, mapping)
            key = clean_key(col[property_row].value, mapping)
            val = clean_val(col[row].value, mapping)
            if not record.get(cat):
                record[cat] = {}
            record[cat][key] = val
        result.append(record)
    return result


if __name__ == "__main__":
    config = ConfigParser()
    config.read(f"{ROOT_DIR}/config/job.conf")
    verbose = config.getboolean("main", "verbose", fallback=False)
    
    bulk_path = config.get("publisher_peer_review", "bulk_path")
    files = glob(f"{bulk_path}/xlsx/*.xlsx")
    data_path = f"{bulk_path}/data"
    os.makedirs(data_path, exist_ok=True)
    
    mapping_file = config.get("publisher_peer_review", "mapping_file")
    with open(mapping_file) as f:
        mapping = json.load(f)
    
    for file in files:
        file_name = os.path.basename(file)
        name, extension = os.path.splitext(file_name)
        records = load_publishers_excel_file(file, mapping)
        print(f"{file}: {len(records)}")
        for n, record in enumerate(records):
            dump_file(f"{data_path}/{name}_{n:03}.json", record, unique=False)
