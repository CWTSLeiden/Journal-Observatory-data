if __name__ == "__main__":
    import sys
    import os
    sys.path.append(os.path.dirname(os.path.dirname(__file__)))

import os
import re
from bulk import dump_file
from datetime import datetime
from glob import glob
from openpyxl import load_workbook


def clean_val(val, mapping={}):
    if type(val) == datetime:
        return str(val.date())
    if not type(val) == str:
        if val == 1:
            return True
        elif val == 0:
            return False
        return val
    if mapping.get(val.strip().lower()):
        return { "@id": mapping.get(val.strip().lower()) }
    if re.match("^https?://", val.strip()):
        return { "@id": val.strip() }
    return val.strip()


def clean_key(key, mapping={}):
    if not key:
        key = "None"
    key = str(key).lower()
    if mapping.get(key): return mapping.get(key)
    key = re.sub("[-/ ]", "_", key)
    key = re.sub("[()]", "", key)
    key = re.sub("_([a-z])", lambda c: c.group(1).upper(), key)
    return key


def load_publishers_excel_file(file, mapping={}):
    provenance_rows = [1, 2, 3]
    category_row = 4
    property_row = 5
    first_data_row = 6
    
    workbook = load_workbook(file)
    worksheet = workbook.active
    result = []
    for row in range(first_data_row, worksheet.max_row):
        record = {}
        # provenance
        for r in provenance_rows:
            key = clean_key(worksheet[r][0].value, mapping)
            val = clean_val(worksheet[r][1].value, mapping)
            if not record.get("provenance"):
                record["provenance"] = {}
            record["provenance"][key] = val
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if col[row].value is None:
                continue
            cat = clean_key(col[category_row].value, mapping)
            key = clean_key(col[property_row].value, mapping)
            val = clean_val(col[row].value, mapping)
            if not record.get(cat):
                record[cat] = {}
            record[cat][key] = val
        result.append(record)
    return result


if __name__ == "__main__":
    from utils import pad_config as config
    
    data_location = config.getpath("publisher_peer_review", "data_location", fallback="data/publisher_peer_review/xlsx")
    data_path = config.getpath("publisher_peer_review", "data_path", fallback="data/publisher_peer_review")
    files = glob(f"{data_location}/*.xlsx")
    data_path = f"{data_path}/data"
    os.makedirs(data_path, exist_ok=True)
    
    mapping = {
        "springer nature": "https://springernature.com",
        "nature": "https://nature.com",
        "wiley": "https://wiley.com",
        "elife sciences publications ltd": "https://elifesciences.org",
        "ieee": "https://www.ieee.org",
        "cc0": "https://creativecommons.org/publicdomain/zero/1.0/",
        "cc-by": "https://creativecommons.org/licenses/by/4.0",
        "cc-by-nc": "https://creativecommons.org/licenses/by-nc/4.0",
        "cc-by-nc-nd": "https://creativecommons.org/licenses/by-nc-nd/4.0",
        "cc-by-nc-sa": "https://creativecommons.org/licenses/by-nc-sa/4.0",
        "cc-by-nd": "https://creativecommons.org/licenses/by-nd/4.0",
        "cc-by-sa": "https://creativecommons.org/licenses/by-sa/4.0",
        "doaj": "https://doaj.org",
        "openalex": "https://openalex.org",
        "sherparomeo": "https://v2.sherpa.ac.uk"
    }
    
    for file in files:
        file_name = os.path.basename(file)
        name, extension = os.path.splitext(file_name)
        records = load_publishers_excel_file(file, mapping)
        print(f"{file}: {len(records)}")
        for n, record in enumerate(records):
            dump_file(f"{data_path}/{name}_{n:03}.json", record, unique=False)
