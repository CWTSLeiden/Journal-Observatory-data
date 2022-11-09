import json
import os
import re
from bulk import dump_file
from configparser import ConfigParser
from datetime import datetime
from glob import glob
from openpyxl import load_workbook
from utils.utils import ROOT_DIR
from convert import dataset_convert


def clean_val(val, mapping={}):
    if type(val) == datetime:
        return str(val.date())
    if not type(val) == str:
        if val == 1:
            return True
        elif val == 0:
            return False
        return val
    if mapping.get(val.lower()):
        return mapping.get(val.lower())
    return val


def clean_key(key, mapping={}):
    if not key:
        key = "None"
    key = str(key)
    key = key.lower()
    if mapping.get(key): return mapping.get(key)
    key = re.sub("[-/ ]", "_", key)
    key = re.sub("[()]", "", key)
    return key


def load_publishers_excel_file(file, mapping={}):
    provenance_rows = [1, 2, 3]
    category_row = 4
    property_row = 5
    first_data_row = 6
    
    workbook = load_workbook(file)
    worksheet = workbook.active
    result = []
    for row in range(first_data_row, worksheet.max_row):
        record = {}
        # provenance
        for r in provenance_rows:
            key = clean_key(worksheet[r][0].value, mapping)
            val = clean_val(worksheet[r][1].value, mapping)
            if not record.get("provenance"):
                record["provenance"] = {}
            record["provenance"][key] = val
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if col[row].value == None:
                continue
            cat = clean_key(col[category_row].value, mapping)
            key = clean_key(col[property_row].value, mapping)
            val = clean_val(col[row].value, mapping)
            if not record.get(cat):
                record[cat] = {}
            record[cat][key] = val
        result.append(record)
    return result


def clean_publisher_record(record):
    result = {}
    result["metadata"] = record.get("metadata")
    result["provenance"] = record.get("provenance")

    ppc = record.get("post_publication_commenting")
    if ppc.get("open"):
        result["post_publication_commenting"] = "open"
    elif ppc.get("on_invitation"):
        result["post_publication_commenting"] = "on_invitation" 

    it = record.get("identity_transparency")
    result["anonymity"] = {}
    result["anonymity"]["author"] = []
    result["anonymity"]["reviewer"] = []
    result["anonymity"]["curator"] = []
    if it.get("triple_anonymised"):
        result["anonymity"]["author"] = [ "curator", "reviewer" ]
        result["anonymity"]["reviewer"] = [ "curator", "author" ]
    if it.get("double_anonymised"):
        result["anonymity"]["author"] = [ "reviewer" ]
        result["anonymity"]["reviewer"] = [ "author" ]
    if it.get("single_anonymised"):
        result["anonymity"]["reviewer"] = [ "author" ]

    riw = record.get("reviewer_interacts_with")
    result["reviewer_interacts_with"] = []
    if riw.get("editor"):
        result["reviewer_interacts_with"].append("editor")
    if riw.get("other_reviewers"):
        result["reviewer_interacts_with"].append("other_reviewers")
    if riw.get("authors"):
        result["reviewer_interacts_with"].append("authors")

    rip = record.get("review_information_published")
    result["review_information_published"] = []
    result["author_opt_in"] = []
    result["reviewer_opt_in"] = []
    if rip.get("review_summaries"):
        result["review_information_published"].append("review_summaries")
    if rip.get("review_reports_author_opt_in"):
        result["review_information_published"].append("review_reports")
        result["author_opt_in"].append("review_reports")
    if rip.get("review_reports_reviewer_opt_in"):
        result["review_information_published"].append("review_reports")
        result["reviewer_opt_in"].append("review_reports")
    if rip.get("review_reports"):
        result["review_information_published"].append("review_reports")
    if rip.get("submitted_manuscripts_author_opt_in"):
        result["review_information_published"].append("submitted_manuscripts")
        result["author_opt_in"].append("submitted_manuscripts")
    elif rip.get("submitted_manuscripts"):
        result["review_information_published"].append("submitted_manuscripts")
    if rip.get("author_editor_communication"):
        result["review_information_published"].append("author_editor_communication")
    if rip.get("reviewer_identities_reviewer_opt_in"):
        result["anonymity"]["reviewer"].append("public")
        result["reviewer_opt_in"].append("identity")
    elif rip.get("reviewer_identities"):
        result["anonymity"]["reviewer"].append("public")
    if rip.get("editor_identities"):
        result["anonymity"]["curator"].append("public")

    return result


if __name__ == "__main__":
    config = ConfigParser()
    config.read(f"{ROOT_DIR}/config/job.conf")
    verbose = config.getboolean("main", "verbose", fallback=False)

    bulk_path = config.get("publisher_peer_review", "bulk_path")
    files = glob(f"{bulk_path}/xlsx/*.xlsx")
    data_path = f"{bulk_path}/data"
    os.makedirs(data_path, exist_ok=True)

    mapping_file = config.get("publisher_peer_review", "mapping_file")
    with open(mapping_file) as f:
        mapping = json.load(f)

    for file in files:
        file_name = os.path.basename(file)
        name, extension = os.path.splitext(file_name)
        records = load_publishers_excel_file(file, mapping)
        print(f"{file}: {len(records)}")
        for n, record in enumerate(records):
            dump_file(f"{data_path}/{name}_{n:03}.json", clean_publisher_record(record), unique=False)

    dataset_convert("publisher_peer_review")
